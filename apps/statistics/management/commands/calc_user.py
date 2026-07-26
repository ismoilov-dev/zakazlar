"""Management command to compute and display statistics for an Employee User ID."""

import os
from decimal import Decimal
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from apps.statistics.services.statistics import StatisticsService


class Command(BaseCommand):
    help = "Calculate employee sales, salary, and performance metrics by User ID."

    def add_arguments(self, parser) -> None:
        parser.add_argument("--user-id", "-u", required=True, type=str, help="Employee User ID (e.g. 0191, 0001)")
        parser.add_argument("--file", "-f", type=str, help="Path to Excel file (e.g. /path/to/workbook.xlsx)")

    def handle(self, *args, **options) -> None:
        user_id = str(options["user_id"]).strip().zfill(4)
        file_path = options.get("file")

        if file_path:
            if not os.path.exists(file_path):
                raise CommandError(f"Excel fayli topilmadi: {file_path}")
            self.stdout.write(self.style.MIGRATE_HEADING(f"📊 Excel faylidan hisob-kitob qilinmoqda: {file_path}"))
            self._calc_from_excel(file_path, user_id)
        else:
            self.stdout.write(self.style.MIGRATE_HEADING(f"📊 Ma'lumotlar bazasidan hisob-kitob qilinmoqda..."))
            try:
                dashboard = StatisticsService().employee_dashboard_for_employee(user_id)
                self._print_dashboard(dashboard)
            except Exception as exc:
                raise CommandError(f"Xato: {exc}")

    def _calc_from_excel(self, file_path: str, user_id: str) -> None:
        wb = load_workbook(file_path, data_only=True)
        if "List1" not in wb.sheetnames or "List2" not in wb.sheetnames:
            raise CommandError("Excel faylida 'List1' va 'List2' sahifalari bo'lishi kerak.")

        s1 = wb["List1"]
        s2 = wb["List2"]

        # Search employee info in List2 first
        emp_info = None
        for r in range(2, s2.max_row + 1):
            uid = s2.cell(row=r, column=2).value
            if uid is not None:
                uid_str = str(uid).strip().zfill(4)
                if uid_str == user_id:
                    emp_info = {
                        "name": str(s2.cell(row=r, column=3).value or "").strip(),
                        "dept": str(s2.cell(row=r, column=4).value or "").strip(),
                    }
                    break

        # Aggregate List1
        total_orders = 0
        successful_orders = 0
        cancelled_orders = 0
        pending_orders = 0
        total_sales = Decimal("0")
        perv_sales = Decimal("0")
        baza_sales = Decimal("0")
        otkaz_sales = Decimal("0")
        v_proc_sales = Decimal("0")
        emp_name = emp_info["name"] if emp_info else ""
        emp_dept = emp_info["dept"] if emp_info else ""

        for r in range(2, s1.max_row + 1):
            uid = s1.cell(row=r, column=8).value
            if uid is None:
                continue
            uid_str = str(uid).strip().zfill(4)
            if uid_str != user_id:
                continue

            if not emp_name:
                emp_name = str(s1.cell(row=r, column=9).value or "").strip()
            if not emp_dept:
                emp_dept = str(s1.cell(row=r, column=10).value or "").strip()

            amount_val = s1.cell(row=r, column=15).value or 0
            try:
                amount = Decimal(str(amount_val)).quantize(Decimal("0.01"))
            except Exception:
                amount = Decimal("0")

            status = str(s1.cell(row=r, column=18).value or "").strip()
            contact = str(s1.cell(row=r, column=20).value or "").strip()
            category = str(s1.cell(row=r, column=22).value or "").strip()

            total_orders += 1
            total_sales += amount

            if status == "Успешно":
                successful_orders += 1
                cat_lower = (category or contact).lower()
                if "baza" in cat_lower or "база" in cat_lower:
                    baza_sales += amount
                else:
                    perv_sales += amount
            elif status == "Отказ":
                cancelled_orders += 1
                otkaz_sales += amount
            elif status in ("В процесс", "У курьера"):
                pending_orders += 1
                v_proc_sales += amount

        if total_orders == 0:
            self.stdout.write(self.style.WARNING(f"⚠️ User ID '{user_id}' bo'yicha hech qanday buyurtma topilmadi."))
            return

        group_upper = emp_dept.upper()
        if group_upper == "BAZA":
            earned_salary = (perv_sales * Decimal("0.12")) + (baza_sales * Decimal("0.12"))
        else:
            earned_salary = (perv_sales * Decimal("0.12")) + (baza_sales * Decimal("0.16"))

        succ_sum = perv_sales + baza_sales
        conv = (float(succ_sum / total_sales) * 100) if total_sales > 0 else 0.0
        denom = total_sales - v_proc_sales
        real_conv = (float(succ_sum / denom) * 100) if denom > 0 else 0.0

        self.stdout.write(self.style.SUCCESS(f"\n✅ USER ID: {user_id} — HISOB-KITOB NATIJALARI"))
        self.stdout.write("=" * 55)
        self.stdout.write(f"👤 Xodim: {emp_name}")
        self.stdout.write(f"🏢 Bo'lim: {emp_dept or 'Noma\'lum'}")
        self.stdout.write("-" * 55)
        self.stdout.write(f"📦 Upakovka (Muvaffaqiyatli): {successful_orders} ta")
        self.stdout.write(f"❌ Otraz (Bekor qilingan):    {cancelled_orders} ta")
        self.stdout.write(f"⏳ V protsess (Jarayonda):    {pending_orders} ta")
        self.stdout.write(f"📋 Jami zakazlar soni:       {total_orders} ta")
        self.stdout.write("-" * 55)
        self.stdout.write(f"💵 Umumiy zakaz summasi:     {total_sales:,.0f} so'm")
        self.stdout.write(f"🎯 Первичный Заказ:          {perv_sales:,.0f} so'm")
        self.stdout.write(f"🔁 База summasi:             {baza_sales:,.0f} so'm")
        self.stdout.write(f"🔴 Отказ summasi:            {otkaz_sales:,.0f} so'm")
        self.stdout.write(f"🟡 В процесс summasi:        {v_proc_sales:,.0f} so'm")
        self.stdout.write("-" * 55)
        self.stdout.write(self.style.SUCCESS(f"💰 ISHLAGAN PULI (Oylik maosh): {earned_salary:,.0f} so'm"))
        self.stdout.write(f"📈 Konversiya: {conv:.2f}% | Real Konversiya: {real_conv:.2f}%")
        self.stdout.write("=" * 55 + "\n")

    def _print_dashboard(self, d) -> None:
        self.stdout.write(self.style.SUCCESS(f"\n✅ USER ID: {d.employee_id} — HISOB-KITOB NATIJALARI"))
        self.stdout.write("=" * 55)
        self.stdout.write(f"👤 Xodim: {d.full_name}")
        self.stdout.write(f"🏢 Bo'lim: {d.group_code or 'Noma\'lum'}")
        self.stdout.write("-" * 55)
        self.stdout.write(f"📦 Upakovka (Muvaffaqiyatli): {d.successful_orders} ta")
        self.stdout.write(f"❌ Otraz (Bekor qilingan):    {d.cancelled_orders} ta")
        self.stdout.write(f"⏳ V protsess (Jarayonda):    {d.pending_orders} ta")
        self.stdout.write(f"📋 Jami zakazlar soni:       {d.total_orders} ta")
        self.stdout.write("-" * 55)
        self.stdout.write(f"💵 Umumiy zakaz summasi:     {d.total_sales:,.0f} so'm")
        self.stdout.write(f"🎯 Первичный Заказ:          {d.perv_sales:,.0f} so'm")
        self.stdout.write(f"🔁 База summasi:             {d.baza_sales:,.0f} so'm")
        self.stdout.write(f"🔴 Отказ summasi:            {d.otkaz_sales:,.0f} so'm")
        self.stdout.write(f"🟡 В процесс summasi:        {d.v_proc_sales:,.0f} so'm")
        self.stdout.write("-" * 55)
        self.stdout.write(self.style.SUCCESS(f"💰 ISHLAGAN PULI (Oylik maosh): {d.earned_salary:,.0f} so'm"))
        self.stdout.write(f"📈 Konversiya: {d.conversion_rate*100:.2f}% | Real Konversiya: {d.real_conversion_rate*100:.2f}%")
        self.stdout.write("=" * 55 + "\n")
