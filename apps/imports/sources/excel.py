"""Excel source parser for uploaded .xlsx workbooks."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import logging

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

from apps.common.services.exceptions import ValidationError
from apps.imports.dto import OrderDTO, PayrollDTO, normalize_employee_id, normalize_order_id
from apps.imports.sources.base import BaseSource
from apps.sales.models import SaleStatus

STATUS_MAP = {
    "успешно": SaleStatus.SUCCESSFUL,
    "отказ": SaleStatus.CANCELLED,
    "в процесс": SaleStatus.PENDING,
    "у курьера": SaleStatus.PENDING,
}


class ExcelSource(BaseSource):
    """Parses `List1` (orders) and `List2` (payroll) from an Excel binary file."""

    def __init__(self, file_content: bytes) -> None:
        self.file_content = file_content

    def read(self) -> tuple[list[OrderDTO], list[PayrollDTO]]:
        try:
            workbook = load_workbook(BytesIO(self.file_content), read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationError(f"Excel faylini ochishda xatolik: {exc}") from exc

        if "List1" not in workbook.sheetnames or "List2" not in workbook.sheetnames:
            raise ValidationError("Faylda List1 va List2 sahifalari bo'lishi shart.")

        orders = self._parse_orders(workbook["List1"])
        payroll = self._parse_payroll(workbook["List2"])

        if not orders:
            raise ValidationError("List1 sahifasida import qilinadigan buyurtmalar topilmadi.")

        return orders, payroll

    def _parse_orders(self, sheet: object) -> list[OrderDTO]:
        headings = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headings:
            raise ValidationError("List1 sahifasi bo'sh.")

        columns = self._columns(
            headings, {"№", "ID", "Ответственный", "Сумма", "Дата Заказа", "статус", "Столбец 2", "Контакт"}
        )
        group_index = next((index for index, value in enumerate(headings) if value == " "), None)
        if group_index is None:
            raise ValidationError("List1 sahifasida guruh ustuni topilmadi.")

        rows: list[OrderDTO] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value is not None for value in values):
                continue
            if values[columns["ID"]] is None or not str(values[columns["ID"]]).strip():
                continue
            try:
                rows.append(
                    OrderDTO(
                        employee_id=normalize_employee_id(values[columns["ID"]]),
                        employee_name=self._text(values[columns["Ответственный"]], "Ответственный"),
                        group_code=self._text(values[group_index], "guruh").upper(),
                        order_id=normalize_order_id(values[columns["№"]]),
                        status=self._status(values[columns["статус"]]),
                        source=self._normalize_source(
                            str(values[columns["Столбец 2"]] or "").strip()
                            or str(values[columns["Контакт"]] or "").strip()
                        ),
                        sale_amount=self._money(values[columns["Сумма"]]),
                        ordered_at=self._datetime(values[columns["Дата Заказа"]]),
                    )
                )
            except ValidationError as exc:
                raise ValidationError(f"List1, {row_number}-qator: {exc}") from exc
        return rows

    def _parse_payroll(self, sheet: object) -> list[PayrollDTO]:
        headings = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headings:
            raise ValidationError("List2 sahifasi bo'sh.")

        columns = self._columns(headings, {"ID", "XODIMLAR ISMLARI ", "Bo'lim ", "OYLIK MOASH "})
        rows: list[PayrollDTO] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in values):
                continue
            employee_id_value = values[columns["ID"]]
            if employee_id_value is None:
                continue
            group_raw = str(values[columns["Bo'lim "]]).strip() if values[columns["Bo'lim "]] is not None else ""
            rows.append(
                PayrollDTO(
                    employee_id=normalize_employee_id(employee_id_value),
                    employee_name=self._text(values[columns["XODIMLAR ISMLARI "]], "XODIMLAR ISMLARI"),
                    group_code=(group_raw or "A").upper(),
                    monthly_salary=self._money(values[columns["OYLIK MOASH "]]),
                )
            )
        return rows

    @staticmethod
    def _columns(headings: tuple[object, ...], required: set[str]) -> dict[str, int]:
        found = {str(value): index for index, value in enumerate(headings) if value is not None}
        missing = required - set(found)
        if missing:
            raise ValidationError(f"Majburiy ustunlar topilmadi: {', '.join(sorted(missing))}")
        return found

    @staticmethod
    def _text(value: object, name: str) -> str:
        result = str(value).strip() if value is not None else ""
        if not result:
            raise ValidationError(f"{name} bo'sh bo'lishi mumkin emas.")
        return result

    @staticmethod
    def _money(value: object, sheet_name: str = "Sheet", row_idx: int | str = "") -> Decimal:
        if value is None:
            return Decimal("0.00")
        s = str(value).strip()
        if not s:
            return Decimal("0.00")
        if s.startswith("#"):
            logger.warning("Excel formula xatosi ('%s') varog': '%s', qator: %s", s, sheet_name, row_idx)
            return Decimal("0.00")
        clean_val = s.replace(" ", "").replace("\xa0", "").replace("$", "").replace("so'm", "").replace("som", "").strip()
        if not clean_val:
            return Decimal("0.00")
        if "." not in clean_val and clean_val.count(",") == 1:
            clean_val = clean_val.replace(",", ".")
        else:
            clean_val = clean_val.replace(",", "")
        try:
            return Decimal(clean_val).quantize(Decimal("0.01"))
        except (InvalidOperation, TypeError, ValueError):
            logger.warning("Noto'g'ri pul summasi formati ('%s') varog': '%s', qator: %s", s, sheet_name, row_idx)
            return Decimal("0.00")

    @staticmethod
    def _status(value: object) -> str:
        raw = str(value).strip().lower() if value is not None else ""
        try:
            return STATUS_MAP[raw]
        except KeyError as exc:
            raise ValidationError(f"Noma'lum status: {value}") from exc

    @staticmethod
    def _datetime(value: object) -> datetime:
        if not isinstance(value, datetime):
            raise ValidationError("Дата Заказа sana-vaqt bo'lishi kerak.")
        return value

    @staticmethod
    def _normalize_source(raw: str) -> str:
        s = raw.strip().lower()
        if "perv" in s or "первич" in s:
            return "Pervichka"
        if "baza" in s or "база" in s:
            return "Baza"
        return raw.strip() if raw.strip() else "Pervichka"
