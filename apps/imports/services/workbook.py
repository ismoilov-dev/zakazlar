"""Importer for the company's operational Excel workbook format."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO

from django.core.files.uploadedfile import UploadedFile
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from apps.common.services.exceptions import ValidationError
from apps.employees.repositories.employee import EmployeeRepository
from apps.groups.repositories.group import SalesGroupRepository
from apps.imports.models import ImportJob, ImportStatus
from apps.imports.repositories.import_job import ImportJobRepository
from apps.sales.models import Sale, SaleStatus
from apps.sales.repositories.sale import SaleRepository

STATUS_MAP = {
    "успешно": SaleStatus.SUCCESSFUL,
    "отказ": SaleStatus.CANCELLED,
    "в процесс": SaleStatus.PENDING,
    "у курьера": SaleStatus.PENDING,
}


@dataclass(frozen=True, slots=True)
class WorkbookRow:
    employee_id: str
    employee_name: str
    group_code: str
    order_id: str
    status: str
    source: str
    sale_amount: Decimal
    ordered_at: datetime


@dataclass(frozen=True, slots=True)
class PayrollRow:
    employee_id: str
    employee_name: str
    group_code: str
    monthly_salary: Decimal


class WorkbookImportService:
    """Import `List1` orders and `List2` employee salaries transactionally."""

    def __init__(self) -> None:
        self.jobs = ImportJobRepository()
        self.groups = SalesGroupRepository()
        self.employees = EmployeeRepository()
        self.sales = SaleRepository()

    def create_job(self, *, workbook: UploadedFile, uploaded_by: object) -> ImportJob:
        if not workbook.name.lower().endswith(".xlsx"):
            raise ValidationError("Faqat .xlsx fayl yuklash mumkin.")
        if workbook.size > 20 * 1024 * 1024:
            raise ValidationError("Excel fayl hajmi 20 MB dan oshmasligi kerak.")
        payload = workbook.read()
        workbook.seek(0)
        checksum = sha256(payload).hexdigest()
        if self.jobs.exists_with_checksum(checksum):
            raise ValidationError("Bu Excel fayl avval import qilingan.")
        return self.jobs.create(workbook=workbook, checksum=checksum, uploaded_by=uploaded_by)

    def process(self, *, job_id: int) -> ImportJob:
        with transaction.atomic():
            job = self.jobs.get_for_processing(job_id)
            if job.status not in {ImportStatus.PENDING, ImportStatus.FAILED}:
                raise ValidationError("Ushbu import qayta ishga tushirilmaydi.")
            self.jobs.mark_processing(job)
        try:
            orders, payroll = self._parse(job)
            with transaction.atomic():
                for row in payroll:
                    group = self.groups.get_or_create(code=row.group_code)
                    self.employees.upsert(
                        employee_id=row.employee_id,
                        full_name=row.employee_name,
                        group=group,
                        monthly_salary=row.monthly_salary,
                    )
                sales: list[Sale] = []
                for row in orders:
                    group = self.groups.get_or_create(code=row.group_code)
                    employee = self.employees.upsert(
                        employee_id=row.employee_id, full_name=row.employee_name, group=group
                    )
                    sales.append(
                        Sale(
                            external_order_id=row.order_id,
                            employee=employee,
                            import_job=job,
                            status=row.status,
                            source=row.source,
                            sale_amount=row.sale_amount,
                            profit_amount=Decimal("0"),
                            ordered_at=row.ordered_at,
                        )
                    )
                created, updated = self.sales.bulk_upsert(sales)
                self.jobs.mark_completed(job, processed=len(orders), created=created, updated=updated)
        except (ValidationError, OSError, ValueError) as exc:
            self.jobs.mark_failed(job, [{"message": str(exc)}])
            raise
        return self.jobs.get(job_id)

    def _parse(self, job: ImportJob) -> tuple[list[WorkbookRow], list[PayrollRow]]:
        job.workbook.open("rb")
        try:
            workbook = load_workbook(BytesIO(job.workbook.read()), read_only=True, data_only=True)
        finally:
            job.workbook.close()
        if "List1" not in workbook.sheetnames or "List2" not in workbook.sheetnames:
            raise ValidationError("Faylda List1 va List2 sahifalari bo'lishi shart.")
        orders = self._parse_orders(workbook["List1"])
        payroll = self._parse_payroll(workbook["List2"])
        if not orders:
            raise ValidationError("List1 sahifasida import qilinadigan buyurtmalar topilmadi.")
        return orders, payroll

    def _parse_orders(self, sheet: object) -> list[WorkbookRow]:
        headings = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headings:
            raise ValidationError("List1 sahifasi bo'sh.")
        columns = self._columns(
            headings, {"№", "ID", "Ответственный", "Сумма", "Дата Заказа", "статус", "Столбец 2", "Контакт"}
        )
        group_index = next((index for index, value in enumerate(headings) if value == " "), None)
        if group_index is None:
            raise ValidationError("List1 sahifasida guruh ustuni topilmadi.")
        rows: list[WorkbookRow] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value is not None for value in values):
                continue
            if values[columns["ID"]] is None or not str(values[columns["ID"]]).strip():
                continue
            try:
                rows.append(
                    WorkbookRow(
                        employee_id=self._employee_id(values[columns["ID"]]),
                        employee_name=self._text(values[columns["Ответственный"]], "Ответственный"),
                        group_code=self._text(values[group_index], "guruh").upper(),
                        order_id=self._text(values[columns["№"]], "№"),
                        status=self._status(values[columns["статус"]]),
                        source=self._normalize_source(
                            str(values[columns["Столбец 2"]] or "").strip()
                            or str(values[columns["Контакт"]] or "").strip()
                        ),
                        sale_amount=self._money(values[columns["Сумма"]], "Сумма"),
                        ordered_at=self._datetime(values[columns["Дата Заказа"]]),
                    )
                )
            except ValidationError as exc:
                raise ValidationError(f"List1, {row_number}-qator: {exc}") from exc
        return rows

    def _parse_payroll(self, sheet: object) -> list[PayrollRow]:
        headings = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headings:
            raise ValidationError("List2 sahifasi bo'sh.")
        columns = self._columns(headings, {"ID", "XODIMLAR ISMLARI ", "Bo'lim ", "OYLIK MOASH "})
        rows: list[PayrollRow] = []
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None for value in values):
                continue
            employee_id_value = values[columns["ID"]]
            if employee_id_value is None:
                continue
            group_raw = str(values[columns["Bo'lim "]]).strip() if values[columns["Bo'lim "]] is not None else ""
            rows.append(
                PayrollRow(
                    employee_id=self._employee_id(employee_id_value),
                    employee_name=self._text(values[columns["XODIMLAR ISMLARI "]], "XODIMLAR ISMLARI"),
                    group_code=(group_raw or "A").upper(),
                    monthly_salary=self._money(values[columns["OYLIK MOASH "]], "OYLIK MOASH"),
                )
            )
        return rows

    @staticmethod
    def _columns(headings: tuple[object, ...], required: set[str]) -> dict[str, int]:
        found = {str(value): index for index, value in enumerate(headings) if value is not None}
        missing = required - set(found)
        if missing:
            raise ValidationError(f"Majburiy ustunlar topilmadi: {', '.join(sorted(missing))}")
        return found

    @staticmethod
    def _text(value: object, name: str) -> str:
        result = str(value).strip() if value is not None else ""
        if not result:
            raise ValidationError(f"{name} bo'sh bo'lishi mumkin emas.")
        return result

    @classmethod
    def _employee_id(cls, value: object) -> str:
        raw = cls._text(value, "ID")
        if raw.endswith(".0"):
            raw = raw[:-2]
        raw = raw.zfill(4)
        if not raw.isdigit():
            raise ValidationError("ID faqat raqamlardan iborat bo'lishi kerak.")
        return raw

    @staticmethod
    def _money(value: object, name: str) -> Decimal:
        if value is None or str(value).strip().startswith("#") or not str(value).strip():
            return Decimal("0.00")
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"))
        except (InvalidOperation, TypeError, ValueError):
            return Decimal("0.00")

    @staticmethod
    def _status(value: object) -> str:
        raw = str(value).strip().lower() if value is not None else ""
        try:
            return STATUS_MAP[raw]
        except KeyError as exc:
            raise ValidationError(f"Noma'lum status: {value}") from exc

    @staticmethod
    def _datetime(value: object) -> datetime:
        if not isinstance(value, datetime):
            raise ValidationError("Дата Заказа sana-vaqt bo'lishi kerak.")
        if timezone.is_naive(value):
            return timezone.make_aware(value, timezone.get_current_timezone())
        return value

    @staticmethod
    def _normalize_source(raw: str) -> str:
        s = raw.strip().lower()
        if "perv" in s or "первич" in s:
            return "Pervichka"
        if "baza" in s or "база" in s:
            return "Baza"
        return raw.strip() if raw.strip() else "Pervichka"

